import configparser
import datetime as dt
import os.path

import openpyxl as xls
import sqlalchemy, sqlalchemy.sql


if __name__ == "__main__":
    cp = configparser.ConfigParser()
    cp.read(os.path.expanduser("~") + "/etc/config.ini")

    gt_engine = sqlalchemy.create_engine(cp.get("DATASOURCES", "gt"))
    gt_metadata = sqlalchemy.MetaData(bind=gt_engine)

    fromDate = dt.date(2020, 11, 1)
    toDate = dt.date(2020, 11, 30)

    gt_conn = gt_engine.connect()

    pedidos = sqlalchemy.Table("pedidos", gt_metadata, autoload=True, schema="gt")
    stmt = pedidos.select().where(sqlalchemy.and_(
        pedidos.c.ide == 1,
        pedidos.c.emisor == "SB_IBERIAN",
        pedidos.c.factura == None,
        pedidos.c.cliente == "37084",
        pedidos.c.fecha_pedido >= fromDate,
        pedidos.c.fecha_pedido <= toDate
    ))

    wb = xls.Workbook()
    ws = wb.active

    i = 0
    result = gt_conn.execute(stmt).fetchall()
    for row in result:
        if i == 0:
            i += 1
            j = 1
            for k in row.keys():
                ws.cell(row=i, column=j, value=str(k))
                j += 1
        i += 1
        j = 1
        for v in row.values():
            ws.cell(row=i, column=j, value=v)
            j += 1






    wb.save(filename="c:/temp/porsche/pedidos.xlsx")
    gt_conn.close()
